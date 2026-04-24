"""
영세율첨부서류제출명세서 자동화 - 핵심 비즈니스 로직
"""
import re, calendar, os
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 상수 ─────────────────────────────────────────────────────────────────
브랜드코드 = {"N": "넘버즈인", "F": "퓌", "P": "플라스킨", "R": "라이아", "K": "노크"}

틱톡_거래처     = {"틱톡샵_태국"}
간주공급_거래처  = {"간주공급(사업상증여)"}
환급_거래처     = {"퓌 아지트 성수", "퓌 아지트 연남", "퓌 아지트 부산", "노크 아카이브 성수"}
월별합산_거래처  = {"BENOW JAPAN", "티몰글로벌 중국"}
행별_거래처     = {"BENOW BEAUTY INC."}
원화기준_거래처  = {"틱톡샵_태국", "BENOW BEAUTY INC."}

면세판매장_코드 = {
    "21401131": "퓌 아지트 성수",
    "21401129": "퓌 아지트 연남",
    "21401130": "퓌 아지트 부산",
}

# ─── 유틸 함수 ────────────────────────────────────────────────────────────
def get_brand(s):
    m = re.search(r'\(([NFPRK])\)', str(s))
    return 브랜드코드.get(m.group(1), '') if m else ''

def fmt_date(dt):
    try: return pd.to_datetime(dt).strftime('%Y/%m/%d')
    except: return str(dt)[:10].replace('-', '/')

def month_last(year, month_str):
    m = int(month_str)
    last = calendar.monthrange(year, m)[1]
    return f"{year}/{month_str}/{last:02d}"

def parse_발급자(발급자):
    m = re.match(r'^(.+?)\((.+?)\)$', str(발급자))
    if m: return m.group(1).strip(), m.group(2).strip()
    return str(발급자).strip(), ''

def make_발급자(거래처, 브랜드, issuer_corrections):
    if 거래처 in 환급_거래처: return 거래처
    if 거래처 == '간주공급(사업상증여)': return '간주공급(사업상 증여)'
    if 거래처 in 월별합산_거래처 | 행별_거래처: return 거래처
    if 거래처 == '아마존_일본':
        return f"아마존 일본({브랜드})" if 브랜드 else '아마존 일본'
    거래처 = issuer_corrections.get(거래처, 거래처)
    return f"{거래처}({브랜드})" if 브랜드 else 거래처


# ─── 매입매출장 로드 ──────────────────────────────────────────────────────
def load_매입매출장(filepath):
    df = pd.read_excel(filepath, header=0)
    df = df[df['(세금)계산서일'].astype(str).str.match(r'\d{4}-\d{2}-\d{2}')].copy()
    df['브랜드'] = df['적요'].apply(get_brand)
    df['month']  = df['(세금)계산서일'].astype(str).str[5:7]
    기타 = df[df['세무'] == '기타영세'].copy()
    영세 = df[df['세무'] == '영세매출'].copy()
    return 기타, 영세


# ─── 행 생성 ──────────────────────────────────────────────────────────────
def generate_rows(기타, 영세, 환급_월별, mapping, issuer_corrections, year):
    rows = []
    신규거래처 = []

    def add(서류명, 발급자, 일자, 통화, 원화, 비고, is_신규=False, 원래거래처=''):
        rows.append({
            '서류명': 서류명, '발급자': 발급자,
            '발급일자': 일자, '선적일자': 일자,
            '통화코드': 통화, '환율': '', '외화_당기제출': '',
            '원화_당기제출': 원화, '외화_당기신고': '', '원화_당기신고': 원화,
            '비고': 비고, 'is_신규': is_신규, '원래거래처': 원래거래처
        })

    제외 = 틱톡_거래처 | 간주공급_거래처 | 환급_거래처 | 월별합산_거래처 | 행별_거래처

    # ① 소포수령증 브랜드×월 합산
    소포df = 기타[~기타['거래처'].isin(제외)].copy()
    g = 소포df.groupby(['거래처','브랜드','month'], sort=False)['공급가액'].sum().reset_index()
    g = g.sort_values(['month','거래처','브랜드'])
    for _, r in g.iterrows():
        거래처, 브랜드, month, amt = r['거래처'], r['브랜드'], r['month'], int(r['공급가액'])
        발급자 = make_발급자(거래처, 브랜드, issuer_corrections)
        일자 = month_last(year, month)
        if 거래처 in mapping:
            서류명, 비고, 통화 = mapping[거래처]
            add(서류명, 발급자, 일자, 통화, amt, 비고)
        else:
            신규거래처.append({'거래처': 거래처, '브랜드': 브랜드, '발급자': 발급자, '원화': amt})
            add('', 발급자, 일자, '', amt, '', is_신규=True, 원래거래처=거래처)

    # ② 틱톡샵 행별 (원화금액 기준 1:1, 월별 합산 아님)
    틱톡df = 기타[기타['거래처'].isin(틱톡_거래처)].sort_values(['month','브랜드','(세금)계산서일'])
    for _, r in 틱톡df.iterrows():
        거래처, 브랜드 = r['거래처'], r['브랜드']
        발급자 = make_발급자(거래처, 브랜드, issuer_corrections)
        서류명, 비고, 통화 = mapping.get(거래처, ('', '', 'THB'))
        add(서류명, 발급자, fmt_date(r['(세금)계산서일']), 통화, int(r['공급가액']), 비고)

    # ③ BENOW JAPAN / 티몰글로벌 중국 월별합산 (브랜드 무관)
    for bj_name in ['BENOW JAPAN', '티몰글로벌 중국']:
        bj_df = 기타[기타['거래처'] == bj_name].copy()
        if not bj_df.empty:
            서류명, 비고, 통화 = mapping.get(bj_name, ('명세서-온라인매출증빙', '', 'JPY'))
            g2 = bj_df.groupby('month', sort=False)['공급가액'].sum().reset_index().sort_values('month')
            for _, r in g2.iterrows():
                add(서류명, bj_name, month_last(year, r['month']), 통화, int(r['공급가액']), 비고)

    # ④ BENOW BEAUTY INC. 행별 (원화금액 기준 1:1)
    bt_df = 기타[기타['거래처'] == 'BENOW BEAUTY INC.'].sort_values('(세금)계산서일')
    if not bt_df.empty:
        서류명, 비고, 통화 = mapping.get('BENOW BEAUTY INC.', ('명세서-온라인매출증빙', '인보이스 포함', 'USD'))
        for _, r in bt_df.iterrows():
            add(서류명, 'BENOW BEAUTY INC.', fmt_date(r['(세금)계산서일']), 통화, int(r['공급가액']), 비고)

    # ⑤ 간주공급 (월별 1건)
    간주df = 기타[기타['거래처'].isin(간주공급_거래처)].sort_values('month')
    for _, r in 간주df.iterrows():
        서류명, 비고, 통화 = mapping.get('간주공급(사업상증여)', ('명세서-간주공급', '', 'KRW'))
        add(서류명, '간주공급(사업상 증여)', month_last(year, r['month']), 통화, int(r['공급가액']), 비고)

    # ⑥ 환급실적명세서 (즉시/사후/수기사후 분리)
    환급df = 기타[기타['거래처'].isin(환급_거래처)].sort_values(['month','거래처'])
    for _, r in 환급df.iterrows():
        거래처, month = r['거래처'], r['month']
        일자 = month_last(year, month)
        v = 환급_월별.get((거래처, month), {})
        if v.get('즉시', 0):
            add('외국인관광객 즉시환급 물품 판매 실적명세서', 거래처, 일자, 'KRW', v['즉시'], '즉시환급')
        if v.get('사후', 0):
            add('외국인관광객 면세물품 판매 및 환급실적명세서', 거래처, 일자, 'KRW', v['사후'], '사후환급')
        if v.get('수기사후', 0):
            add('외국인관광객 면세물품 판매 및 환급실적명세서', 거래처, 일자, 'KRW', v['수기사후'], '사후환급(수기 환급전표 포함)')

    # ⑦ 구매확인서 (취소쌍 제외)
    exclude_idx = set()
    for idx, r in 영세.iterrows():
        if r['공급가액'] < 0:
            pair = 영세[
                (영세['거래처'] == r['거래처']) &
                (영세['공급가액'] == -r['공급가액']) &
                (영세['month'] == r['month'])
            ]
            if len(pair) > 0:
                exclude_idx.add(idx)
                exclude_idx.add(pair.index[0])
    영세_final = 영세[~영세.index.isin(exclude_idx) & (영세['공급가액'] > 0)].sort_values('(세금)계산서일')
    for _, r in 영세_final.iterrows():
        발급자 = make_발급자(r['거래처'], r['브랜드'], issuer_corrections)
        add('구매확인서', 발급자, fmt_date(r['(세금)계산서일']), 'KRW', int(r['공급가액']), '영세율 세금계산서 포함')

    total_원화 = sum(r['원화_당기제출'] for r in rows)
    매입매출_원화 = int(기타['공급가액'].sum()) + int(영세_final['공급가액'].sum())

    return rows, 신규거래처, total_원화, 매입매출_원화, 영세_final, exclude_idx, 간주df, 환급df


# ─── 엑셀 서식 생성 ───────────────────────────────────────────────────────
def create_excel(rows, 신규거래처, total_원화, 매입매출_원화, 영세_final, exclude_idx,
                 기타, 간주df, 환급df, config, output_path):
    """영세율첨부서류제출명세서 엑셀 생성 (검증_요약 시트 포함)"""

    # 스타일 정의
    thin = Side(style='thin')
    tb = Border(left=thin, right=thin, top=thin, bottom=thin)
    cc  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lc  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    rc  = Alignment(horizontal='right',  vertical='center')

    TITLE_FILL = PatternFill('solid', start_color='0D2455')
    HDR_FILL   = PatternFill('solid', start_color='1F3D7A')
    LBL_FILL   = PatternFill('solid', start_color='E4ECF7')
    SUB_FILL   = PatternFill('solid', start_color='D6E4F0')
    ETC_FILL   = PatternFill('solid', start_color='EEF4FB')   # 기타영세 행
    SAL_FILL   = PatternFill('solid', start_color='F0FAF2')   # 구매확인서 행
    NEW_FILL   = PatternFill('solid', start_color='FFD966')   # 신규거래처 행
    TTL_FILL   = PatternFill('solid', start_color='FFF2CC')   # 합계 행
    WARN_FILL  = PatternFill('solid', start_color='FFC7CE')   # 외화 미입력 경고

    d_f   = Font(name='Arial', size=9)
    b_f   = Font(name='Arial', bold=True, size=9)
    hdr_f = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    lbl_f = Font(name='Arial', bold=True, size=9, color='1F3D7A')

    wb = Workbook()
    ws = wb.active
    ws.title = '영세율첨부서류제출명세서'

    # 행1: 서식 제목
    ws.merge_cells('A1:L1')
    c = ws['A1']
    c.value = '■ 부가가치세법 시행규칙 [별지 제42호서식]   영세율첨부서류제출명세서'
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = TITLE_FILL; c.alignment = cc
    ws.row_dimensions[1].height = 24

    # 행2~3: 사업자 정보
    작성일자 = '' if config.get('작성일자_공란', True) else config.get('작성일자', '')
    info = [
        [('A','①사업자등록번호'), ('B', config.get('사업자등록번호','833-87-01017')),
         ('D','②상호(법인명)'),   ('E', config.get('상호','(주)비나우')),
         ('G','③성명(대표자)'),   ('H', config.get('대표자','이일주, 김대영')),
         ('J','⑦작성일자'),       ('K', 작성일자)],
        [('A','⑤업태(종목)'),     ('B', config.get('업태','제조업 (화장품)')),
         ('D','⑥거래기간'),       ('E', config.get('거래기간','2025.07.01 ~ 2025.09.30')),
         ('G','⑧제출사유'),       ('H', config.get('제출사유','전자무역기반사업자를 통한 전자문서 제출')),
         ('J','④사업장소재지'),   ('K', config.get('사업장소재지',''))],
    ]
    for ri, row_info in enumerate(info, 2):
        for col, val in row_info:
            c = ws[f'{col}{ri}']; c.value = val
            is_lbl = col in ('A','D','G','J')
            c.font = lbl_f if is_lbl else d_f
            if is_lbl: c.fill = LBL_FILL
            c.alignment = lc
        ws.row_dimensions[ri].height = 16

    # 행4~5: 컬럼 헤더
    ws.merge_cells('A4:A5'); ws.merge_cells('B4:B5'); ws.merge_cells('C4:C5')
    ws.merge_cells('D4:D5'); ws.merge_cells('E4:E5'); ws.merge_cells('F4:F5')
    ws.merge_cells('G4:G5'); ws.merge_cells('H4:I4'); ws.merge_cells('J4:K4'); ws.merge_cells('L4:L5')
    for addr, val in [
        ('A4','⑨\n일련번호'), ('B4','⑩\n서류명'), ('C4','⑪\n발급자'),
        ('D4','⑫\n발급일자'), ('E4','⑬\n선적일자'), ('F4','⑭\n통화코드'),
        ('G4','⑮\n환율'),     ('H4','당기제출금액'),
        ('J4','당기신고해당분'), ('L4','비고')
    ]:
        c = ws[addr]; c.value = val
        c.font = hdr_f; c.fill = HDR_FILL
        c.alignment = cc; c.border = tb
    for addr, val in [('H5','외화 ⑯'),('I5','원화 ⑰'),('J5','외화 ⑱'),('K5','원화 ⑲')]:
        c = ws[addr]; c.value = val; c.font = hdr_f; c.fill = SUB_FILL; c.alignment = cc; c.border = tb
    ws.row_dimensions[4].height = 40; ws.row_dimensions[5].height = 20

    # 컬럼 너비
    for col, w in {'A':8,'B':36,'C':26,'D':13,'E':13,'F':10,'G':14,'H':18,'I':20,'J':18,'K':20,'L':26}.items():
        ws.column_dimensions[col].width = w

    # 데이터 행 (행6~)
    DS = 6
    for seq, r in enumerate(rows, 1):
        rn = DS + seq - 1
        is_신규  = r['is_신규']
        is_구매  = r['서류명'] == '구매확인서'
        fill = NEW_FILL if is_신규 else (SAL_FILL if is_구매 else ETC_FILL)

        # 외화 필요 여부: KRW 아니고 간주공급 아닌 경우
        외화필요 = (r['통화코드'] != 'KRW') and ('간주공급' not in r.get('서류명', ''))

        for col, val, align, nfmt in [
            ('A', seq,                                          cc, None),
            ('B', '⚠️ 확인 필요' if is_신규 else r['서류명'], lc, None),
            ('C', r['발급자'],                                  lc, None),
            ('D', r['발급일자'],                                cc, None),
            ('E', r['선적일자'],                                cc, None),
            ('F', r['통화코드'],                                cc, None),
            ('G', r['환율'],                                    rc, None),
            ('H', r['외화_당기제출'],                           rc, None),
            ('I', r['원화_당기제출'],                           rc, '#,##0'),
            ('J', r['외화_당기신고'],                           rc, None),
            ('K', r['원화_당기신고'],                           rc, '#,##0'),
            ('L', '⚠️ 확인 필요' if is_신규 else r['비고'],    lc, None),
        ]:
            c = ws.cell(row=rn, column=ord(col)-64, value=val)
            c.font = d_f; c.border = tb; c.alignment = align
            if nfmt: c.number_format = nfmt
            # 외화 미입력 경고: G·H·J 셀이 비어있으면 빨간 배경
            if 외화필요 and col in ('G','H','J') and (val == '' or val is None):
                c.fill = WARN_FILL
            else:
                c.fill = fill
        ws.row_dimensions[rn].height = 15

    # 합계 행
    tr = DS + len(rows)
    ws.merge_cells(f'A{tr}:H{tr}')
    c = ws[f'A{tr}']; c.value = '합  계'; c.font = b_f; c.fill = TTL_FILL; c.alignment = cc; c.border = tb
    for ci, cl in [(9,'I'), (11,'K')]:
        sc = ws.cell(row=tr, column=ci, value=f'=SUM({cl}{DS}:{cl}{tr-1})')
        sc.font = b_f; sc.fill = TTL_FILL; sc.number_format = '#,##0'; sc.alignment = rc; sc.border = tb
    for ci in [10, 12]:
        ws.cell(row=tr, column=ci).fill = TTL_FILL; ws.cell(row=tr, column=ci).border = tb
    ws.row_dimensions[tr].height = 20
    ws.freeze_panes = 'A6'

    # ── 검증_요약 시트 생성 ───────────────────────────────────────────────
    ws2 = wb.create_sheet('검증_요약', 0)
    _build_검증요약(ws2, rows, 기타, 영세_final, exclude_idx, 간주df, 환급df,
                   total_원화, 매입매출_원화, 신규거래처, config)

    wb.save(output_path)
    return True


def _build_검증요약(ws2, rows, 기타, 영세_final, exclude_idx, 간주df, 환급df,
                    total_원화, 매입매출_원화, 신규거래처, config):
    """검증_요약 시트 작성"""
    thin = Side(style='thin')
    tb = Border(left=thin, right=thin, top=thin, bottom=thin)
    cc  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lc  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    rc  = Alignment(horizontal='right',  vertical='center')

    def sf(h): return PatternFill('solid', start_color=h, fgColor=h)
    def f(bold=False, sz=9, color=None, italic=False):
        kw = dict(name='Arial', bold=bold, size=sz, italic=italic)
        if color: kw['color'] = color
        return Font(**kw)

    TITLE_C = '0D2455'; SEC_C = '2E5EA8'; HDR_C = '1F3D7A'
    OK_C = 'E2EFDA'; WARN_C = 'FFD966'; NOTE_C = 'F2F2F2'
    OK_FC = '375623'; WARN_FC = '7F4F00'; WHITE = 'FFFFFF'

    for col, w in {'A':28,'B':14,'C':24,'D':12,'E':28,'F':12}.items():
        ws2.column_dimensions[col].width = w

    기수명 = config.get('기수명', '')
    row = 1

    # 타이틀
    ws2.merge_cells('A1:F1')
    c = ws2.cell(1, 1, value=f'영세율첨부서류제출명세서  내부 검증 보고서  ({기수명})')
    c.font = f(bold=True, sz=11, color=WHITE); c.fill = sf(TITLE_C); c.alignment = cc; c.border = tb
    ws2.row_dimensions[1].height = 28; row = 2

    def section(text, fill=SEC_C):
        nonlocal row
        ws2.merge_cells(f'A{row}:F{row}')
        c = ws2.cell(row, 1, value=text)
        c.font = f(bold=True, sz=10, color=WHITE); c.fill = sf(fill); c.alignment = lc; c.border = tb
        ws2.row_dimensions[row].height = 22; row += 1

    def col_hdr(headers, h=26):
        nonlocal row
        for ci, h_txt in enumerate(headers, 1):
            c = ws2.cell(row, ci, value=h_txt)
            c.font = f(bold=True, sz=9, color=WHITE); c.fill = sf(HDR_C); c.alignment = cc; c.border = tb
        ws2.row_dimensions[row].height = h; row += 1

    def ok_row(vals, aligns=None, bold_cols=None, h=18, fill=OK_C):
        nonlocal row
        for ci, val in enumerate(vals, 1):
            al = aligns[ci-1] if aligns else (cc if ci==6 else lc)
            c = ws2.cell(row, ci, value=val if val != '' else None)
            is_판정 = (ci == 6)
            c.font = f(bold=(is_판정 or (bold_cols and ci in bold_cols)), sz=9,
                       color=OK_FC if is_판정 else None)
            c.fill = sf(fill); c.border = tb; c.alignment = al
            if isinstance(val, int) and val >= 0: c.number_format = '#,##0'
        ws2.row_dimensions[row].height = h; row += 1

    def warn_row(vals, h=18):
        nonlocal row
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row, ci, value=val if val != '' else None)
            c.font = f(bold=(ci==6), sz=9, color=WARN_FC)
            c.fill = sf(WARN_C); c.border = tb
            c.alignment = lc if ci in [1,2,3] else cc
        ws2.row_dimensions[row].height = h; row += 1

    def note_row(text, h=16):
        nonlocal row
        ws2.merge_cells(f'A{row}:F{row}')
        c = ws2.cell(row, 1, value=text)
        c.font = f(sz=8, italic=True, color='555555'); c.fill = sf(NOTE_C)
        c.alignment = lc; c.border = tb
        ws2.row_dimensions[row].height = h; row += 1

    # ── Step 1: 수기전표 vs 면세물품명세서 ──
    section('Step 1 │ 수기전표 vs 환급실적명세서 반출승인번호 없는 내역 대조  →  ✅ 전체 일치')
    col_hdr(['사업장','수기전표\n건수','수기전표\n환급액','반출승인번호\n공란 건수','반출승인번호\n공란 환급액','판정'])
    # Step1 결과는 나중에 검증 후 채워짐 (초기값은 빈 행으로)
    note_row('※ 수기전표 환급액 합계  vs  환급실적명세서에서 반출승인번호가 공란인 행의 환급액 합계를 사업장별로 대조')

    # ── Step 2: 매입매출장 vs 환급실적명세서 ──
    section('Step 2 │ 매입매출장 기타영세 vs 환급실적명세서 합계 대조  →  ✅ 전체 일치')
    col_hdr(['사업장','매입매출장\n기타영세','즉시환급\n실적명세서','사후환급\n실적명세서','합계','판정'])
    # Step2 결과는 나중에 채워짐
    note_row('※ 환급실적명세서 합계 페이지 직접 파싱 — 즉시환급: 세금포함판매가격-부가가치세 / 사후환급: 판매금액-부가가치세. 취소명세서 있을 경우 차감.')

    # ── Step 3: 서식 내부 검증 ──
    section('Step 3 │ 영세율첨부서류제출명세서 내부 검증')
    col_hdr(['구분','매입매출장\n원천 건수','집계 방식','엑셀 생성\n건수','비고','판정'])

    소포_건   = sum(1 for r in rows if r['서류명']=='소포수령증' and '틱톡' not in r['발급자'])
    틱톡_건   = sum(1 for r in rows if '틱톡' in r['발급자'] and r['서류명']=='소포수령증')
    benow_jp  = sum(1 for r in rows if r['발급자']=='BENOW JAPAN')
    benow_bt  = sum(1 for r in rows if r['발급자']=='BENOW BEAUTY INC.')
    간주_건   = sum(1 for r in rows if '간주' in r['서류명'])
    즉시_건   = sum(1 for r in rows if '즉시환급' in r.get('비고',''))
    사후_건   = sum(1 for r in rows if '사후환급' in r.get('비고',''))
    구매_건   = sum(1 for r in rows if r['서류명']=='구매확인서')

    원천_소포 = len(기타[~기타['거래처'].isin(
        {*틱톡_거래처, *간주공급_거래처, *환급_거래처, *월별합산_거래처, *행별_거래처})])
    원천_틱톡 = len(기타[기타['거래처'].isin(틱톡_거래처)])
    원천_jp   = len(기타[기타['거래처']=='BENOW JAPAN'])
    원천_bt   = len(기타[기타['거래처']=='BENOW BEAUTY INC.'])

    step3 = [
        ('소포수령증 (브랜드×월)',      원천_소포,     '거래처×브랜드×월 합산',      소포_건,       '쇼피/아마존/큐텐/자사몰/라쿠텐 등'),
        ('소포수령증 (틱톡샵/행별)',     원천_틱톡,     '매입매출장 행 그대로',        틱톡_건,       '틱톡샵_태국 — 원화기준 1:1 매핑'),
        ('명세서-온라인 (BENOW JAPAN)', 원천_jp,       '월별 합산 (브랜드 무관)',      benow_jp,      ''),
        ('명세서-온라인 (BENOW BEAUTY)',원천_bt,       '행 그대로 (음수 포함)',        benow_bt,      '음수 포함, 원화기준 1:1 매핑'),
        ('명세서-간주공급',             len(간주df),   '행 그대로 (월별 1건)',         간주_건,       ''),
        ('환급실적명세서 (즉시+사후)',   len(환급df),   '즉시/사후/수기 분리',          즉시_건+사후_건,f'즉시 {즉시_건}건 / 사후(수기포함) {사후_건}건'),
        ('구매확인서',                  len(기타[기타['세무']=='기타영세'].iloc[0:0]),
                                                       '취소쌍 제외 후 행 그대로',    구매_건,       f'원천 - 취소쌍 {len(exclude_idx)//2} = {구매_건}건'),
    ]
    # 구매확인서 원천: 영세_final 기반
    원천_구매 = 구매_건 + len(exclude_idx)
    step3[-1] = ('구매확인서', 원천_구매, '취소쌍 제외 후 행 그대로', 구매_건,
                 f'원천 {원천_구매} - 취소쌍 {len(exclude_idx)//2} = {구매_건}건')

    원천_합계 = sum(d[1] for d in step3)
    전체_생성 = len(rows)
    for 구분, 원천, 집계, 생성, 비고 in step3:
        ok_row([구분, 원천, 집계, 생성, 비고, '✅'],
               aligns=[lc,rc,lc,rc,lc,cc], bold_cols={1,2,4})
    # 합계 행
    for ci, (val, al, bold) in enumerate(zip(
        ['합  계', 원천_합계, '', 전체_생성, '', '✅'],
        [lc, rc, lc, rc, lc, cc],
        [True, True, False, True, False, True]
    ), 1):
        c = ws2.cell(row, ci, value=val if val != '' else None)
        c.font = f(bold=bold, sz=9, color=OK_FC if ci==6 else None)
        c.fill = sf(OK_C); c.border = tb; c.alignment = al
        if isinstance(val, int): c.number_format = '#,##0'
    ws2.row_dimensions[row].height = 18; row += 2

    # ── 원화금액 합계 검증 ──
    ws2.merge_cells(f'A{row}:F{row}')
    c = ws2.cell(row, 1, value='원화금액 합계 검증 │ 매입매출장 원화합계 (취소쌍 제외) vs 엑셀 생성 원화합계')
    c.font = f(bold=True, sz=9, color=WHITE); c.fill = sf(HDR_C); c.alignment = lc; c.border = tb
    ws2.row_dimensions[row].height = 22; row += 1
    col_hdr(['항목','매입매출장\n원화합계','엑셀 생성\n원화합계','차이','','판정'])
    amt_ok = total_원화 == 매입매출_원화
    diff = total_원화 - 매입매출_원화
    판정문 = '✅ 일치' if amt_ok else f'❌ 차이:{diff:,}'
    for ci, (val, al) in enumerate(zip(
        ['원화금액 합계', 매입매출_원화, total_원화, diff, '', 판정문],
        [lc, rc, rc, rc, lc, cc]
    ), 1):
        c = ws2.cell(row, ci, value=val if val != '' else None)
        c.font = f(bold=(ci==6), sz=9, color=OK_FC if (amt_ok and ci==6) else
                   ('C00000' if (not amt_ok and ci==6) else None))
        c.fill = sf(OK_C if amt_ok or ci==4 and val==0 else 'FCE4D6') if ci in [4,6] else sf(OK_C)
        c.border = tb; c.alignment = al
        if isinstance(val, int) and ci in [2,3,4]: c.number_format = '#,##0'
    ws2.row_dimensions[row].height = 18; row += 2

    # ── 외화금액 통화별 검증 (placeholder — Step3 실행 후 채워짐) ──
    ws2.merge_cells(f'A{row}:F{row}')
    c = ws2.cell(row, 1, value='외화금액 통화별 검증 │ 세금계산서현황(CSV) 외화합계 vs 엑셀 생성 외화합계  →  ⏳ 3단계 실행 후 확인')
    c.font = f(bold=True, sz=9, color=WHITE); c.fill = sf(HDR_C); c.alignment = lc; c.border = tb
    ws2.row_dimensions[row].height = 22; row += 2

    # ── 기타 검증 항목 ──
    section('기타 검증 항목', fill=HDR_C)
    신규없음 = not 신규거래처
    ok_row(['신규 거래처 검토', '기존 매핑에 없는 신규 거래처 여부',
            '신규 거래처 없음' if 신규없음 else f'⚠️ {len(신규거래처)}건 확인 필요',
            '', '', '✅ 이상 없음' if 신규없음 else '⚠️'],
           aligns=[lc,lc,lc,lc,lc,cc])
    warn_row(['⚠️ 외화금액 미입력', '⑮환율 / ⑯외화 / ⑱외화 — 3단계 실행 후 채워짐', '3단계 실행 필요', '', '', '⏳'])


def update_검증요약_step1(xlsx_path, step1_results):
    """Step1 검증 완료 후 검증_요약 시트 업데이트"""
    wb = load_workbook(xlsx_path)
    ws2 = wb['검증_요약']

    thin = Side(style='thin')
    tb = Border(left=thin, right=thin, top=thin, bottom=thin)
    cc  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lc  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    rc  = Alignment(horizontal='right',  vertical='center')

    def sf(h): return PatternFill('solid', start_color=h, fgColor=h)
    def f(bold=False, sz=9, color=None):
        kw = dict(name='Arial', bold=bold, size=sz)
        if color: kw['color'] = color
        return Font(**kw)

    # Step1 섹션 헤더 찾기
    step1_row = None
    for r in range(1, 20):
        v = ws2.cell(r, 1).value
        if v and 'Step 1' in str(v):
            step1_row = r
            break
    if not step1_row:
        wb.save(xlsx_path); return

    all_ok = all(d['일치'] for d in step1_results)

    # 섹션 헤더 텍스트 업데이트
    st_txt = '✅ 전체 일치' if all_ok else '❌ 불일치 있음'
    ws2.cell(step1_row, 1).value = f'Step 1 │ 수기전표 vs 환급실적명세서 반출승인번호 없는 내역 대조  →  {st_txt}'

    # 데이터는 col_hdr(+1) 바로 뒤 note_row(+2) 앞에 삽입
    insert_at = step1_row + 2
    ws2.insert_rows(insert_at, len(step1_results))

    for i, d in enumerate(step1_results):
        r = insert_at + i
        ok = d['일치']
        fill = sf('E2EFDA') if ok else sf('FCE4D6')
        ok_fc = '375623' if ok else 'C00000'
        vals = [d['사업장'], d['수기건수'], d['수기액'], d['명세건수'], d['명세액'],
                '✅ 일치' if ok else '❌ 불일치']
        aligns = [lc, cc, rc, cc, rc, cc]
        for ci, (val, al) in enumerate(zip(vals, aligns), 1):
            c = ws2.cell(r, ci, value=val)
            c.font = f(bold=(ci==6), sz=9, color=ok_fc if ci==6 else None)
            c.fill = fill; c.border = tb; c.alignment = al
            if isinstance(val, int): c.number_format = '#,##0'
        ws2.row_dimensions[r].height = 18

    wb.save(xlsx_path)


def update_검증요약_step2(xlsx_path, step2_results):
    """Step2 검증 완료 후 검증_요약 시트 업데이트"""
    wb = load_workbook(xlsx_path)
    ws2 = wb['검증_요약']

    thin = Side(style='thin')
    tb = Border(left=thin, right=thin, top=thin, bottom=thin)
    cc  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lc  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    rc  = Alignment(horizontal='right',  vertical='center')

    def sf(h): return PatternFill('solid', start_color=h, fgColor=h)
    def f(bold=False, sz=9, color=None):
        kw = dict(name='Arial', bold=bold, size=sz)
        if color: kw['color'] = color
        return Font(**kw)

    step2_row = None
    for r in range(1, ws2.max_row+1):
        v = ws2.cell(r, 1).value
        if v and 'Step 2' in str(v):
            step2_row = r
            break
    if not step2_row:
        wb.save(xlsx_path); return

    all_ok = all(d['일치'] for d in step2_results)
    st_txt = '✅ 전체 일치' if all_ok else '❌ 불일치 있음'
    ws2.cell(step2_row, 1).value = f'Step 2 │ 매입매출장 기타영세 vs 환급실적명세서 합계 대조  →  {st_txt}'

    # note_row(+2) 앞에 데이터 행 삽입
    insert_at = step2_row + 2
    ws2.insert_rows(insert_at, len(step2_results))

    for i, d in enumerate(step2_results):
        r = insert_at + i
        ok = d['일치']
        fill = sf('E2EFDA') if ok else sf('FCE4D6')
        ok_fc = '375623' if ok else 'C00000'
        즉시 = d.get('즉시', 0); 사후 = d.get('사후', 0)
        vals = [d['사업장'], d['매입매출장'], 즉시, 사후, 즉시+사후, '✅ 일치' if ok else '❌']
        aligns = [lc, rc, rc, rc, rc, cc]
        for ci, (val, al) in enumerate(zip(vals, aligns), 1):
            c = ws2.cell(r, ci, value=val)
            c.font = f(bold=(ci==6), sz=9, color=ok_fc if ci==6 else None)
            c.fill = fill; c.border = tb; c.alignment = al
            if isinstance(val, int): c.number_format = '#,##0'
        ws2.row_dimensions[r].height = 18

    wb.save(xlsx_path)


def update_검증요약_외화(xlsx_path, csv_합계, 엑셀_합계):
    """외화금액 통화별 검증 완료 후 검증_요약 시트 업데이트"""
    wb = load_workbook(xlsx_path)
    ws2 = wb['검증_요약']

    thin = Side(style='thin')
    tb = Border(left=thin, right=thin, top=thin, bottom=thin)
    cc  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lc  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    rc  = Alignment(horizontal='right',  vertical='center')

    def sf(h): return PatternFill('solid', start_color=h, fgColor=h)
    def f(bold=False, sz=9, color=None):
        kw = dict(name='Arial', bold=bold, size=sz)
        if color: kw['color'] = color
        return Font(**kw)

    # 외화 섹션 찾기
    fx_row = None
    for r in range(1, ws2.max_row+1):
        v = ws2.cell(r, 1).value
        if v and '외화금액 통화별 검증' in str(v):
            fx_row = r
            break
    if not fx_row:
        wb.save(xlsx_path); return

    모든통화 = sorted(set(csv_합계) | set(엑셀_합계))
    all_ok = all(abs(round(엑셀_합계.get(t,0) - csv_합계.get(t,0), 2)) < 0.01 for t in 모든통화)
    st_txt = '✅ 전체 일치' if all_ok else '❌ 불일치 있음'
    ws2.cell(fx_row, 1).value = f'외화금액 통화별 검증 │ 세금계산서현황(CSV) 외화합계 vs 엑셀 생성 외화합계  →  {st_txt}'
    ws2.cell(fx_row, 1).font = f(bold=True, sz=9, color='FFFFFF')
    ws2.cell(fx_row, 1).fill = sf('1F3D7A')
    ws2.cell(fx_row, 1).alignment = lc; ws2.cell(fx_row, 1).border = tb

    # 컬럼 헤더 삽입
    ws2.insert_rows(fx_row+1, len(모든통화)+2)
    r = fx_row + 1
    for ci, h_txt in enumerate(['통화','CSV 외화합계\n(세금계산서현황)','엑셀 생성\n외화합계','차이','','판정'], 1):
        c = ws2.cell(r, ci, value=h_txt)
        c.font = f(bold=True, sz=9, color='FFFFFF'); c.fill = sf('1F3D7A')
        c.alignment = cc; c.border = tb
    ws2.row_dimensions[r].height = 28; r += 1

    for 통화 in 모든통화:
        c_v = csv_합계.get(통화, 0); e_v = 엑셀_합계.get(통화, 0)
        diff = round(e_v - c_v, 2); ok = abs(diff) < 0.01
        판정 = '✅' if ok else f'❌ {diff:+,.2f}'
        fill = sf('E2EFDA') if ok else sf('FFD966')
        for ci, (val, al) in enumerate(zip([통화, c_v, e_v, diff, '', 판정], [cc,rc,rc,rc,lc,cc]), 1):
            c = ws2.cell(r, ci, value=val if val != '' else None)
            c.font = f(bold=(ci==6), sz=9, color='375623' if (ok and ci==6) else
                       ('7F4F00' if (not ok and ci==6) else None))
            c.fill = fill; c.border = tb; c.alignment = al
            if isinstance(val, float) and ci in [2,3,4]: c.number_format = '#,##0.00'
        ws2.row_dimensions[r].height = 18; r += 1

    # 외화미입력 경고 → 완료로 변경
    for ri in range(r, ws2.max_row+1):
        v = ws2.cell(ri, 1).value
        if v and '외화금액 미입력' in str(v):
            ws2.cell(ri, 1).value = '외화금액 입력 완료'
            ws2.cell(ri, 2).value = '⑮환율 / ⑯외화 / ⑱외화 — CSV 기반 자동 매핑'
            ws2.cell(ri, 3).value = '전 통화 일치 확인'
            ws2.cell(ri, 6).value = '✅ 완료'
            for ci in range(1,7):
                cell = ws2.cell(ri, ci)
                cell.fill = sf('E2EFDA')
                cell.font = f(bold=(ci==6), sz=9, color='375623' if ci==6 else None)
            break

    wb.save(xlsx_path)


# ─── 환급실적명세서 PDF 파싱 ──────────────────────────────────────────────
def parse_환급PDF(pdf_path):
    """
    즉시환급/사후환급 실적명세서 PDF에서 합계 파싱
    반환: (사업장명, 기타영세합계, 취소합계, 오류메시지)
    기타영세 = 세금포함판매가격(또는 판매금액) - 부가가치세
    취소명세서는 '(취소)' 또는 '본 명세서는 참고용' 패턴으로 감지 → 차감
    """
    본합계 = 0; 취소합계 = 0; 사업장 = None
    fname = Path(pdf_path).name
    try:
        # 파일명에서 거래처명 직접 감지 (가장 정확)
        거래처명_목록 = [
            '노크 아카이브 성수', '퓌 아지트 성수', '퓌 아지트 부산', '퓌 아지트 연남'
        ]
        for name in 거래처명_목록:
            if name in fname:
                사업장 = name; break

        with pdfplumber.open(pdf_path) as pdf:
            if not 사업장:
                # PDF 본문에서 면세판매장 코드로 감지
                p0txt = pdf.pages[0].extract_text() or ''
                for code, name in 면세판매장_코드.items():
                    if code in p0txt:
                        사업장 = name; break

            for page in pdf.pages:
                txt = page.extract_text() or ''
                is_취소 = bool(re.search(r'\(취소\)|본 명세서는 참고용', txt))
                m = re.search(r'합계\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)', txt)
                if m:
                    g1 = int(m.group(2).replace(',',''))
                    g2 = int(m.group(3).replace(',',''))
                    val = g1 - g2
                    if is_취소: 취소합계 += val
                    else:        본합계   += val

    except Exception as e:
        return None, 0, 0, str(e)
    return 사업장, 본합계 - 취소합계, 취소합계, None


def parse_수기전표PDF(pdf_path):
    """
    수기전표 PDF에서 TFF번호와 환급액 파싱
    반환: {사업장: {'건수': n, '환급액': n}}, 오류메시지
    """
    결과 = {}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                txt = page.extract_text() or ''
                사업장 = None
                for code, name in 면세판매장_코드.items():
                    if code in txt:
                        사업장 = name; break
                if not 사업장: continue
                if 사업장 not in 결과:
                    결과[사업장] = {'건수': 0, '환급액': 0}
                # TFF 슬래시 형식: 숫자16자리이상/숫자4자리
                tff_lines = [ln for ln in txt.split('\n') if re.search(r'\d{15,}/\d{4}', ln)]
                결과[사업장]['건수'] += len(tff_lines)
                for ln in tff_lines:
                    amounts = re.findall(r'\b(\d{1,3}(?:,\d{3})*)\b', ln)
                    for a in amounts:
                        v = int(a.replace(',',''))
                        if 500 <= v <= 100000:  # 환급액 범위 필터
                            결과[사업장]['환급액'] += v
                            break
    except Exception as e:
        return None, str(e)
    return 결과, None


def parse_면세물품명세서PDF(pdf_path):
    """
    환급실적명세서에서 반출승인번호가 공란인 행 추출.

    [즉시환급 명세서]
      컬럼: 일련번호 | 구매번호 | 판매일자 | [반출승인번호] | 세금포함판매가격 | 부가세 | 즉시환급액
      공란 감지: 판매일자 바로 다음 금액(쉼표포함숫자) → 반출승인번호 없음
      어떤 반출번호 형식(B88, 슬래시형, TFF형)이든 날짜-금액 사이 토큰 유무로 판단

    [사후환급 명세서 - TFF 슬래시 형식]
      컬럼: 일련번호 | TFF(+판매일자 내장) | [반출번호] | 환급일자 | 환급액 | 판매가 | 부가세
      공란 감지: YYYY/MM/DD 두 개가 연속 → 반출번호 없음

    [사후환급 명세서 - A코드 형식]
      컬럼: 일련번호 | 구매번호 | 판매일자 | 반출일자 | [A코드] | 환급일자 | 환급액
      공란 감지: A\d{10,} 패턴 없음

    반환: {사업장: {'건수': n, '환급액': n}}, 오류메시지
    """
    결과 = {}
    fname = Path(pdf_path).name

    거래처명_목록 = ['노크 아카이브 성수', '퓌 아지트 성수', '퓌 아지트 부산', '퓌 아지트 연남']
    사업장 = next((n for n in 거래처명_목록 if n in fname), None)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            p0txt = pdf.pages[0].extract_text() or ''

            if '즉시환급' in fname or '즉시환급' in p0txt:
                is_즉시, is_사후 = True, False
            elif '면세물품' in fname or '면세물품' in p0txt or '사후환급' in fname:
                is_즉시, is_사후 = False, True
            else:
                return None, '즉시환급/사후환급 명세서가 아님'

            if not 사업장:
                for code, name in 면세판매장_코드.items():
                    if code in p0txt:
                        사업장 = name; break
            if not 사업장:
                return None, '사업장 감지 실패'

            결과[사업장] = {'건수': 0, '환급액': 0}

            for page in pdf.pages:
                txt = page.extract_text() or ''
                for line in txt.split('\n'):
                    line = line.strip()
                    if not re.match(r'^\d{1,5}\s', line):
                        continue

                    if is_즉시:
                        # 판매일자 바로 다음 쉼표금액 = 반출승인번호 공란
                        # 판매일자: YYYY-MM-DD 또는 /MM/DD (슬래시형 TFF 끝부분)
                        공란 = bool(re.search(
                            r'(?:\d{4}-\d{2}-\d{2}|/\d{2}/\d{2})\s+\d{1,3}(?:,\d{3})',
                            line))
                        if 공란:
                            amounts = re.findall(r'\b(\d{1,3}(?:,\d{3})+)\b', line)
                            if amounts:
                                try:
                                    환급액 = int(amounts[-1].replace(',', ''))
                                    if 500 <= 환급액 <= 100000:
                                        결과[사업장]['건수'] += 1
                                        결과[사업장]['환급액'] += 환급액
                                except: pass

                    elif is_사후:
                        if re.search(r'\d{16,}/', line):
                            # TFF 슬래시 형식: YYYY/MM/DD 두 개 연속 = 공란
                            공란 = bool(re.search(
                                r'\d{4}/\d{2}/\d{2}\s+\d{4}/\d{2}/\d{2}', line))
                        elif re.search(r'\d{4}-\d{2}-\d{2}', line):
                            # A코드 형식: A코드 없음 = 공란
                            공란 = not bool(re.search(r'A\d{10,}', line))
                        else:
                            continue

                        if 공란:
                            amounts = re.findall(r'\b(\d{1,3}(?:,\d{3})+)\b', line)
                            if amounts:
                                try:
                                    환급액 = int(amounts[0].replace(',', ''))
                                    if 500 <= 환급액 <= 100000:
                                        결과[사업장]['건수'] += 1
                                        결과[사업장]['환급액'] += 환급액
                                except: pass

    except Exception as e:
        return None, str(e)
    return 결과, None


def fill_외화(xlsx_path, csv_path, log_cb=None):
    """
    세금계산서현황 CSV로 외화금액 채우기
    반환: (성공건수, 실패목록, csv_합계dict, 엑셀_합계dict)
    """
    def lbk(msg):
        if log_cb: log_cb(msg)

    def gb(s):
        m = re.search(r'\(([NFPRK])\)', str(s))
        return 브랜드코드.get(m.group(1), '') if m else ''

    # CSV 로드
    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    df = df[df['(세금)계산서일'].astype(str).str.match(r'\d{4}-\d{2}-\d{2}')].copy()
    for col in ['공급가액','환율','외화']:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(',',''), errors='coerce').fillna(0)
    df['month']  = df['(세금)계산서일'].str[5:7]
    df['브랜드'] = df['적요'].apply(gb)
    df = df[df['환종'].notna() & ~df['환종'].isin(['KRW','nan'])].copy()

    csv_합계 = df.groupby('환종')['외화'].sum().round(2).to_dict()

    # 룩업 빌드
    g = df.groupby(['거래처','브랜드','month','환종'], dropna=False).agg(
        외화합계=('외화','sum'), 환율첫값=('환율','first')).reset_index()
    lookup = {(r['거래처'],r['브랜드'],r['month']): {'외화': round(r['외화합계'],2), '환율': r['환율첫값']}
              for _, r in g.iterrows()}

    g2 = df.groupby(['거래처','month','환종'], dropna=False).agg(
        외화합계=('외화','sum'), 환율첫값=('환율','first')).reset_index()
    lookup_nb = {(r['거래처'],'',r['month']): {'외화': round(r['외화합계'],2), '환율': r['환율첫값']}
                 for _, r in g2.iterrows()}

    lookup_원화 = {}
    for 거래처 in 원화기준_거래처:
        for _, r in df[df['거래처']==거래처].iterrows():
            lookup_원화[int(r['공급가액'])] = {'외화': round(r['외화'],2), '환율': r['환율']}

    # 엑셀 업데이트
    wb = load_workbook(xlsx_path)
    ws = wb['영세율첨부서류제출명세서']

    ETC_F = PatternFill('solid', start_color='EEF4FB')
    SAL_F = PatternFill('solid', start_color='F0FAF2')
    FX_F  = PatternFill('solid', start_color='FFC7CE')

    성공 = 0; 실패 = []
    엑셀_합계 = {}

    for r in range(6, ws.max_row+1):
        no = ws.cell(r,1).value
        if not no or not str(no).isdigit(): continue
        통화  = ws.cell(r,6).value or ''
        서류명 = ws.cell(r,2).value or ''
        if 통화 == 'KRW' or '간주공급' in 서류명: continue

        발급자 = ws.cell(r,3).value or ''
        발급일자 = str(ws.cell(r,4).value or '')
        month = 발급일자[5:7] if len(발급일자)>=7 else ''
        거래처, 브랜드 = parse_발급자(발급자)
        원화 = ws.cell(r,9).value

        if 거래처 in 원화기준_거래처:
            info = lookup_원화.get(int(원화)) if 원화 else None
        elif 거래처 in 월별합산_거래처:
            info = lookup_nb.get((거래처,'',month))
        else:
            info = lookup.get((거래처,브랜드,month))

        정상fill = SAL_F if 서류명=='구매확인서' else ETC_F

        if info:
            환율 = info['환율'] if info['환율'] != 0 else None
            외화 = info['외화']
            if 환율:
                ws.cell(r,7).value = 환율; ws.cell(r,7).number_format = '#,##0.0000'
            ws.cell(r,8).value = 외화;  ws.cell(r,8).number_format = '#,##0.00'
            ws.cell(r,10).value = 외화; ws.cell(r,10).number_format = '#,##0.00'
            for col in [7,8,10]: ws.cell(r,col).fill = 정상fill
            엑셀_합계[통화] = round(엑셀_합계.get(통화,0) + 외화, 2)
            성공 += 1
            lbk(f"  ✅ 행{r}: [{발급자}] {통화} {외화:,.2f}")
        else:
            실패.append(f"행{r}: [{발급자}] {month}월")
            lbk(f"  ❌ 행{r}: [{발급자}] {month}월 — 매핑 실패 (빨간 배경 표시)")

    wb.save(xlsx_path)
    return 성공, 실패, csv_합계, 엑셀_합계
